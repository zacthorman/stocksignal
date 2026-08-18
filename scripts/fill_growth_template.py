"""Build a filled ZipTraderU Growth Template workbook for one ticker.

    PYTHONPATH=src python scripts/fill_growth_template.py --input examples/DEMO.json

The output mirrors the original template's three tabs and every one of its cell
formulas, with three deliberate additions: a reasoning column on the scorecard,
a band lookup so the growth adjustment stops being a number retyped by hand, and
a legend naming exactly which cells to edit.

WHY THIS WRITES FORMULAS AND NOT COMPUTED NUMBERS. `growth_template.py` can work
the whole chain out in Python, and a workbook of its answers would be a dead
report. The point of the spreadsheet is that the scorecard is a judgement that
changes: you read one more analyst piece, move Competitive Advantage from 18 to
22, and the total crosses from 165 to 170, which flips the band from +0% to +10%,
which moves the fair multiple and every price target. That chain has to be live
in the file or the workbook lies the moment anyone edits it.

WHY THE BAND IS A LOOKUP RATHER THAN A TYPED NUMBER. In the original, cell B5 is
typed in by hand off a printed table, so nothing stops a 165 total sitting next
to a +50% adjustment. Same for the projected growth rates in C8:C11, which are
manual and not wired to the anticipated rate computed one sheet over. Both are
places where the file can silently disagree with itself, and both are wired up
here.

READ `growth_template.py`'s MODULE DOCSTRING BEFORE USING ANY NUMBER THIS
PRODUCES. None of this method appears in the 256-page course; it is a Discord
handout with no worked example and nine of its twelve inputs unguided, and its
fair multiple is linear and uncapped. It is a second opinion on a different
axis from the chart, which is exactly why it is worth having, and it is not a
valuation anyone should defend as such.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from stocksignal.growth_template import (  # noqa: E402
    BANDS,
    CATEGORIES,
    MATURITY_MAX,
    SP500_GROWTH,
    SP500_MULTIPLE,
)

COMPANY = "Company"
QUALITATIVE = "Qualitative Factors"
CALC = "Calculation Center"

FONT = "Arial"
INPUT_FONT = Font(name=FONT, color="0000FF")
"""Blue: a hardcoded input. The financial-model convention, and it is the only
thing telling a reader which numbers are evidence and which are derived."""
FORMULA_FONT = Font(name=FONT, color="000000")
LINK_FONT = Font(name=FONT, color="008000")
"""Green: pulled from another sheet."""
HEADER = Font(name=FONT, bold=True)
FILL_ME = PatternFill("solid", fgColor="FFFF00")
NOTE = Font(name=FONT, italic=True, size=9)

MONEY = "$#,##0;($#,##0);-"
PCT = "0.0%"
MULTIPLE = "0.0x"


def build(spec: dict, out: Path) -> Path:
    ticker = spec["ticker"]
    years = spec["years"]
    revenues = spec["revenues"]
    shares = spec["shares"]
    scores = spec.get("scores", {})
    projection_years = spec.get("projection_years", 4)

    if not (len(years) == len(revenues) == len(shares)):
        raise SystemExit("years, revenues and shares must be the same length")

    wb = Workbook()
    calc = wb.active
    calc.title = CALC
    company = wb.create_sheet(COMPANY)
    qual = wb.create_sheet(QUALITATIVE)

    _build_qualitative(qual, scores)
    _build_calc(calc, len(years))
    _build_company(company, ticker, years, revenues, shares, projection_years)

    wb.save(out)
    return out


def _build_qualitative(ws, scores: dict) -> None:
    """The twelve categories, verbatim, plus a column the original does not have.

    The reasoning column is the single most useful change to this sheet. Nine of
    these categories carry no scoring rule at all, so the number on its own
    records a mood. Six months later "Competitive Advantage: 22" is unreadable,
    and "22, only vendor with FDA clearance in the category, two competitors
    withdrew in 2025" can be argued with. The band that these totals select
    swings the final price target by a factor of six, which is far too much
    leverage to rest on numbers nobody has to justify.
    """
    ws["A1"] = "Qualitative Factors"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)

    for column, heading, width in (
        ("A", "Factor", 62),
        ("B", "Points", 9),
        ("C", "Out of", 9),
        ("D", "Reasoning (required: nine of these have no scoring rule)", 70),
    ):
        ws[f"{column}2"] = heading
        ws[f"{column}2"].font = HEADER
        ws.column_dimensions[column].width = width

    row = 3
    for key, label, maximum in CATEGORIES:
        entry = scores.get(key, {})
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name=FONT)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{row}"] = entry.get("points", 0)
        ws[f"B{row}"].font = INPUT_FONT
        ws[f"B{row}"].fill = FILL_ME
        ws[f"C{row}"] = f"/{maximum}"
        ws[f"D{row}"] = entry.get("reasoning", "")
        ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    maturity = scores.get("maturity", {})
    ws[f"A{row}"] = (
        "Maturity PENALTY: has accelerated revenue growth started slowing? "
        "(this SUBTRACTS from the total)"
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"B{row}"] = maturity.get("points", 0)
    ws[f"B{row}"].font = INPUT_FONT
    ws[f"B{row}"].fill = FILL_ME
    ws[f"C{row}"] = f"/({MATURITY_MAX})"
    ws[f"D{row}"] = maturity.get("reasoning", "")
    ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    maturity_row = row

    total_row = row + 1
    ws[f"A{total_row}"] = "Total"
    ws[f"A{total_row}"].font = HEADER
    # Matches the original's `=SUM(B3:B13)-B14`: the eleven positives, minus the
    # maturity penalty.
    ws[f"B{total_row}"] = f"=SUM(B3:B{maturity_row - 1})-B{maturity_row}"
    ws[f"B{total_row}"].font = Font(name=FONT, bold=True)
    ws[f"C{total_row}"] = "/200"
    ws[f"B{total_row}"].comment = Comment(
        "The eleven positive categories cap at exactly 200. Maturity subtracts up to 25, "
        "so the real range is -25 to 200, not 0 to 200.",
        "stocksignal",
    )

    legend = total_row + 2
    ws[f"A{legend}"] = "HOW TO USE THIS SHEET"
    ws[f"A{legend}"].font = HEADER
    for offset, line in enumerate(
        (
            "Edit the yellow cells in column B only. Everything else is derived.",
            "Column D is not optional. Only three of these twelve categories have a scoring",
            "rule in the template (cash runway, market share, track record). The other nine are",
            "unguided, and the band this total lands in multiplies the growth rate by anything",
            "from 1.5 down to 0.25, so an unexplained score moves the price target by a factor",
            "of six with nothing behind it.",
            "",
            "Scoring rules the template does give:",
            "  Company Health: 1yr of cash = 10, 2yr = 15, 3yr = 20, 5yr = 25",
            "  Market Share: 10% = 2, 25% = 5, 50% = 10, 75%+ = 15",
            "  Track Record: 1yr consistent growth = 5, 3yr = 10, 5yr = 15",
        ),
        start=1,
    ):
        ws[f"A{legend + offset}"] = line
        ws[f"A{legend + offset}"].font = NOTE


def _build_calc(ws, history_years: int) -> None:
    """The Calculation Center, with the band lookup the original leaves manual."""
    ws["A1"] = "Calculation Center"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 14
    for column in ("D", "E", "F"):
        ws.column_dimensions[column].width = 16

    growth_rows = history_years - 1
    first, last = 4, 3 + growth_rows

    ws["A4"] = f"Average revenue growth rate (mean of {growth_rows} historical years)"
    ws["B4"] = f"=AVERAGE('{COMPANY}'!C{first}:C{last})"
    ws["B4"].font = LINK_FONT
    ws["B4"].number_format = PCT

    ws["A5"] = "Scorecard total"
    ws["B5"] = f"='{QUALITATIVE}'!B{len(CATEGORIES) + 4}"
    ws["B5"].font = LINK_FONT

    ws["A6"] = "Anticipated growth rate change (from the band table, right)"
    # MATCH with match_type 1 needs the lookup column ascending, which is why
    # the band table below is written low-to-high rather than in the template's
    # printed order. IFERROR catches a total below the lowest band, where the
    # template says to abandon this model entirely rather than pick a number.
    ws["B6"] = (
        f"=IFERROR(INDEX($E$4:$E${3 + len(BANDS)},MATCH(B5,$D$4:$D${3 + len(BANDS)},1)),"
        '"BELOW 120: use an earnings model, or avoid")'
    )
    ws["B6"].number_format = PCT
    ws["B6"].comment = Comment(
        "In the original template this cell is typed in by hand off a printed table, so "
        "nothing stops a 165 total sitting next to a +50% adjustment. Wired to the table "
        "here so the two cannot disagree.",
        "stocksignal",
    )

    ws["A7"] = "Band"
    ws["B7"] = (
        f'=IFERROR(INDEX($F$4:$F${3 + len(BANDS)},MATCH(B5,$D$4:$D${3 + len(BANDS)},1)),"n/a")'
    )

    ws["A9"] = "Anticipated revenue growth rate"
    ws["B9"] = '=IF(ISNUMBER(B6),B4+(B4*B6),"")'
    ws["B9"].number_format = PCT
    ws["B9"].font = Font(name=FONT, bold=True)

    ws["A10"] = "S&P 500 average, last 10 years"
    ws["B10"] = SP500_GROWTH
    ws["B10"].font = INPUT_FONT
    ws["B10"].number_format = PCT
    ws["B10"].comment = Comment(
        "From the original template, cell B10, labelled 'S&P 500 Average Last 10 Years'. It "
        "sits under a revenue growth rate and yields a price-to-sales multiple, so it has to "
        "mean the index's average SALES growth. The template never says, and the course never "
        "mentions the number. Unverified.",
        "stocksignal",
    )

    ws["A11"] = "Average S&P 500 multiple, last 10 years"
    ws["B11"] = SP500_MULTIPLE
    ws["B11"].font = INPUT_FONT
    ws["B11"].number_format = MULTIPLE
    ws["B11"].comment = Comment(
        "From the original template, cell B11. Unsourced and undated.", "stocksignal"
    )

    ws["A12"] = "Your fair multiple = (anticipated growth / S&P average) x S&P multiple"
    ws["B12"] = (
        '=IF(ISNUMBER(B9),IF(B9<=0,"NEGATIVE GROWTH: model does not apply",(B9/B10)*B11),"")'
    )
    ws["B12"].number_format = MULTIPLE
    ws["B12"].font = Font(name=FONT, bold=True)
    ws["B12"].comment = Comment(
        "Linear and uncapped. 35% growth earns 15x sales; 70% earns 30x. Markets do not "
        "sustain that, so treat a high multiple as the model running away rather than as a "
        "finding. Above roughly 25x, stop using the number.",
        "stocksignal",
    )

    ws["D2"] = "Band table (from the template's rating system)"
    ws["D2"].font = HEADER
    for column, heading in (("D", "Score from"), ("E", "Adjustment"), ("F", "Band")):
        ws[f"{column}3"] = heading
        ws[f"{column}3"].font = HEADER
    # Ascending, because MATCH(...,1) requires it.
    for offset, (low, _high, adjustment, label) in enumerate(reversed(BANDS)):
        row = 4 + offset
        ws[f"D{row}"] = low
        ws[f"E{row}"] = adjustment
        ws[f"E{row}"].number_format = PCT
        ws[f"F{row}"] = label
        ws[f"F{row}"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["F"].width = 44


def _build_company(
    ws, ticker: str, years: list[int], revenues: list[float], shares: list[float], ahead: int
) -> None:
    """Revenue, shares and the price target per year, with the original's formulas."""
    ws["A1"] = f"{ticker}"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)

    headings = (
        ("A", "Year", 10),
        ("B", "Revenue ($)", 18),
        ("C", "Growth Rate", 13),
        ("D", "Shares Outstanding (incl. dilution)", 30),
        ("E", "Fair Value Multiple (P/S)", 22),
        ("F", "Price Target", 14),
        ("G", "Source", 44),
    )
    for column, heading, width in headings:
        ws[f"{column}2"] = heading
        ws[f"{column}2"].font = HEADER
        ws.column_dimensions[column].width = width

    row = 3
    first_history = row
    for index, (year, revenue, share_count) in enumerate(zip(years, revenues, shares)):  # noqa: B905
        ws[f"A{row}"] = str(year)
        ws[f"B{row}"] = revenue
        ws[f"B{row}"].font = INPUT_FONT
        ws[f"B{row}"].fill = FILL_ME
        ws[f"B{row}"].number_format = MONEY
        if index > 0:
            ws[f"C{row}"] = f'=IF(B{row - 1}=0,"",(B{row}-B{row - 1})/B{row - 1})'
            ws[f"C{row}"].number_format = PCT
        ws[f"D{row}"] = share_count
        ws[f"D{row}"].font = INPUT_FONT
        ws[f"D{row}"].fill = FILL_ME
        ws[f"D{row}"].number_format = "#,##0"
        _target_cells(ws, row)
        ws[f"G{row}"] = "Reported. Cite the 10-K or 10-Q."
        ws[f"G{row}"].font = NOTE
        row += 1

    last_history = row - 1
    first_projection = row
    for step in range(ahead):
        ws[f"A{row}"] = str(years[-1] + step + 1)
        # Wired to the anticipated rate rather than typed in. In the original,
        # C8:C11 are manual and disconnected from the Calculation Center, so a
        # workbook could compute 40% one sheet over and project 5% here.
        ws[f"C{row}"] = f"='{CALC}'!$B$9"
        ws[f"C{row}"].number_format = PCT
        ws[f"C{row}"].font = LINK_FONT
        ws[f"B{row}"] = f'=IF(ISNUMBER(C{row}),B{row - 1}*(1+C{row}),"")'
        ws[f"B{row}"].number_format = MONEY
        ws[f"D{row}"] = f"=D{last_history}"
        ws[f"D{row}"].number_format = "#,##0"
        _target_cells(ws, row)
        ws[f"G{row}"] = "Projected. Share count held flat: dilution NOT modelled."
        ws[f"G{row}"].font = NOTE
        row += 1

    notes = row + 1
    ws[f"A{notes}"] = "HOW TO USE THIS SHEET"
    ws[f"A{notes}"].font = HEADER
    for offset, line in enumerate(
        (
            f"Edit the yellow cells only: revenue and share count for {years[0]} to {years[-1]}.",
            f"Rows {first_history} to {last_history} are reported history. Rows {first_projection} "
            "onwards are projected and derive from the Calculation Center.",
            "",
            "Price Target = (Revenue / Shares Outstanding) x Fair Value Multiple, which is the",
            "original template's formula. It is a price-to-sales valuation, so it ignores costs,",
            "margins, debt and cash entirely: two companies with identical revenue value the same",
            "whether one is profitable and the other is burning cash.",
            "",
            "Share count is held flat across the projection, so dilution is modelled as zero. The",
            "column header asks for dilution to be included. On the kind of high-beta names this",
            "scans, check recent 424B5 and S-3 filings before trusting the later years.",
            "",
            "None of this method is in the 256-page course. It is a Discord handout with no worked",
            "example. Treat it as a second opinion on a different axis from the chart, never as",
            "confirmation of it.",
        ),
        start=1,
    ):
        ws[f"A{notes + offset}"] = line
        ws[f"A{notes + offset}"].font = NOTE


def _target_cells(ws, row: int) -> None:
    """Columns E and F, identical on every row. The original's `=(B/D)*E`.

    Guarded against a zero share count, which the original divides by happily.
    """
    ws[f"E{row}"] = f"='{CALC}'!$B$12"
    ws[f"E{row}"].number_format = MULTIPLE
    ws[f"E{row}"].font = LINK_FONT
    ws[f"F{row}"] = (
        f'=IF(OR(D{row}=0,NOT(ISNUMBER(E{row})),NOT(ISNUMBER(B{row}))),"",(B{row}/D{row})*E{row})'
    )
    ws[f"F{row}"].number_format = "$#,##0.00"
    ws[f"F{row}"].font = Font(name=FONT, bold=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="JSON spec for one ticker.")
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    spec = json.loads(args.input.read_text())
    out = args.out or Path(f"out/Growth Template - {spec['ticker']}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    build(spec, out)
    print(f"written {out}")
    print("now run the xlsx skill's recalc.py over it: formulas have no cached values yet")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
